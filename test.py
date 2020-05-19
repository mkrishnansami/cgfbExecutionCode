# Lets read the Trade Approval file
# Then create another output file and append trade execution data to it
import os
import openpyxl



TradeDate = "05122020"
NameofFile = "TradeApprovalCapitalForTradeDate"+ TradeDate+ ".xlsx"
a = os.getcwd()

CopyString = "copy H:" + "\\" + NameofFile + " " + a
print(CopyString)
os.system(CopyString)
#
print(NameofFile)
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(NameofFile)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active


print(sheet_obj.max_row)
print(sheet_obj.max_column)
RowNum = sheet_obj.max_row
ColNum = sheet_obj.max_column

l = 0
m = 0

ValueList = list()

for i in range(10,RowNum+1):
    TempList = list()
    for j in range(3,13):
        cell_obj = sheet_obj.cell(row=i, column=j)
        TempList.append(cell_obj.value)
        #print(cell_obj.value)
    ValueList.append(TempList)

print(ValueList)

# Now read in execution info from execution file
NameofFile = 'Trades.xlsx'
wb_obj = openpyxl.load_workbook(NameofFile)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
RowNum = sheet_obj.max_row
ColNum = sheet_obj.max_column

TradeValueList = list()
TradeIsinList = list()
for i in range(2,RowNum+1):
    TempList = list()
    Notional = None
    for j in range(1,ColNum+1):
        cell_obj = sheet_obj.cell(row=i, column=j)
        if j==5:
            #Counterparty
            TempList.append(cell_obj.value)
        elif j==8:
            TradeIsinList.append(cell_obj.value)
        elif j==11:
            #Nominal
            Notional = cell_obj.value
            TempList.append(cell_obj.value*1000)
            #Yield
            TempList.append(" ")
        elif j==12:
            #Price
            TempList.append(cell_obj.value)
        elif j==14:
            TempList.append(cell_obj.value)




    TradeValueList.append(TempList)


print(TradeValueList)
print(TradeIsinList)
print(len(ValueList))
print(len(TradeValueList))

# Now open the template Trade Execution file and write the ValueList and TradeValueList into that file
# Lets make a copy of the TradeExecutionTemplate.xlsx file
NameofFile = "OrderForm-SAMI" + TradeDate + "_TradeExecutions.xlsx"
SystemString = "copy TradeExecutionTemplate.xlsx " + NameofFile
print(SystemString)
os.system(SystemString)

# Now write data
wb_obj = openpyxl.load_workbook(NameofFile)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

for i in range(10, 10+len(ValueList)):
    cell1 = sheet_obj.cell(i,20)
    cell1.value = 2
    for j in range(3,13):
        cell = sheet_obj.cell(i,j)
        cell.value = ValueList[i-10][j-3]


wb_obj.save(NameofFile)


# Now also print trade execution data
wb_obj = openpyxl.load_workbook(NameofFile)
sheet_obj = wb_obj.active

# Write trade date: Cell B7
C1 = sheet_obj["B7"]
C1.value = "Trade Date: " + TradeDate


for i in range(10, 10+len(ValueList)):
    ValueIsinTemp = sheet_obj.cell(i,3).value
    cell1 = sheet_obj.cell(i, 20)
    cell1.value = ' '
    print("ValueIsinTemp = ", ValueIsinTemp)
    for k in range(0, len(TradeValueList)):
        if ValueIsinTemp == TradeIsinList[k]:
            for j in range(15, 20):
                cell = sheet_obj.cell(i, j)
                cell.value = TradeValueList[k][j - 15]
                if j==16:
                    if sheet_obj.cell(i,7).value != TradeValueList[k][j - 15]:
                        cell1 = sheet_obj.cell(i, 20)
                        cell1.value = 1



wb_obj.save(NameofFile)
